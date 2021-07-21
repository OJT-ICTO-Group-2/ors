# -*- coding: utf-8 -*-
# -------------------------------------------------------------------------
# This is a sample controller
# this file is released under public domain and you can use without limitations
# -------------------------------------------------------------------------
def make_ordinal(n):
    n = int(n)
    suffix = ['th', 'st', 'nd', 'rd', 'th'][min(n % 10, 4)]
    if 11 <= (n % 100) <= 13:
        suffix = 'th'
    return str(n) + suffix


def index():
    students = db(db.student).select(orderby=db.student.id)
    return locals()

def tor_download():
    # function to generate an xlsx file of the student's TOR

    student_id = request.args(0)
    if not student_id:
        raise HTTP(400, "Bad request")

    # get data of student from database
    student = db(db.student.student_id == student_id).select().first()
    college = db(db.college.id == student.college_id).select().first()
    program = db(db.program.id == student.program_id).select().first()

    if student.specialization_id:
        specialization = db(db.specialization.id == student.specialization_id).select().first()
    else:
        specialization = None

    grades = db(db.grade.student_id == student.id).select(join=db.course.on(db.grade.course_id == db.course.id),
                                                          orderby=db.grade.term_year | db.grade.term_sem)

    terms = set()       # set of terms; terms are of type tuple: (term_sem, term_year); to be used as keys for term_data dictionary

    # store to terms set all terms attended by student
    for grade in grades:
        terms.add((grade.grade.term_sem, grade.grade.term_year))

    # sort the terms by school year
    terms = sorted(terms, key=lambda tup: tup[0])
    terms = sorted(terms, key=lambda tup: tup[1])

    term_grades = {}

    # iterate over the terms
    for term in terms:
        # initialize the contents of term_data
        term_grades[term] = []

        # check for each grade and store to term_data[term] if matches the term
        for grade in grades:
            if grade.grade.term_sem == term[0] and grade.grade.term_year == term[1]:
                term_grades[term].append(grade)

    transcript = db(db.transcript.student_id == student.id).select().first()

    attestor = db(db.faculty.id == transcript.attestor_id).select().first()     # indorsed by
    reviewer = db(db.faculty.id == transcript.reviewer_id).select().first()     # reviewed by
    registrar = db(db.staff.id == transcript.registrar_id).select().first()     # certified true

    # script for generating xlsx file of TOR
    import openpyxl as opx

    # store the file name
    file_name = f'{student.last_name}-{student.first_name}-{student.middle_name}-TOR.xlsx'.replace(" ", "_")

    # load template file
    wb = opx.load_workbook(filename='applications/ors/static/templates/TOR-Template.xlsx')

    # get current worksheet
    ws = wb.active

    # write student data in header
    ws['I9'] = f'{student.last_name}, {student.first_name} {student.middle_name}'.upper()
    ws['I10'] = student.address
    ws['J11'] = student.birthdate.strftime("%B %-d, %Y")
    ws['J12'] = student.birthplace
    ws['K13'] = (" ").join(program.name.split(" ")[:4])
    ws['K14'] = (" ").join(program.name.split(" ")[4:])
    ws['K15'] = specialization.name if specialization else ""
    try:
        ws['J16'] = transcript.date_conferred.strftime("%B %-d, %Y")
    except:
        pass
    ws['J17'] = f'per BOR Referendum No. {transcript.ref_no}, s. {transcript.series}'

    ws['E14'] = student.last_attended
    ws['E15'] = student.category
    ws['E16'] = college.name
    ws['E17'] = student.sem_admitted

    # initialize counters
    i = 20
    j = 20
    row_counter = 0
    page_count = 1

    # iterate for each term and write grades in the file
    for term in terms:
        if term[0] == "Summer":
            ws[f'A{j}'] = f'{term[0]}, {term[1]}'
        elif len(term_grades[term]) < 2:
            ws[f'A{j}'] = f'{term[0].replace("Semester", "Sem.")}, {term[1]}'
        else:
            ws[f'A{j}'] = term[0]
            if row_counter + 1 >= 23: ws[f'A{j+44}'] = term[1]
            else: ws[f'A{j+1}'] = term[1]
        j += len(term_grades[term])

        for grade in term_grades[term]:
            # course code
            ws.merge_cells(f'C{i}:D{i}')
            ws[f'C{i}'] = grade.course.code

            #course title
            ws[f'E{i}'] = grade.course.title

            # final grade
            if grade.grade.final_grade == 4:
                ws[f'L{i}'] = "Inc."
            elif grade.grade.final_grade == 0:
                ws[f'L{i}'] = "Drp."
            else:
                ws[f'L{i}'] = grade.grade.final_grade

            # removal rating
            ws[f'M{i}'] = grade.grade.removal_rating

            # course units
            if grade.grade.final_grade == 4 and grade.grade.removal_rating is None or grade.grade.final_grade == 0:
                ws[f'N{i}'] = "-"
            else:
                ws[f'N{i}'] = grade.course.units

            # increment counters
            row_counter += 1
            i += 1

            # if one page is filled up, move to the next page
            if row_counter >= 23:
                i += 43
                j += 43
                row_counter = 0
                page_count += 1

    # if student graduated already, write the graduation details
    if student.date_graduated:
        for x in range(3):
            ws.unmerge_cells(f'E{i + x}:K{i + x}')
        # ws.merge_cells(start_row=i, start_column=1, end_row=i+2, end_column=14)
        ws.merge_cells(f'A{i}:N{i+2}')
        ws[f'A{i}'] = f'''GRADUATED WITH THE DEGREE OF {program.name.upper()} ({program.abbreviation}) ON
                          {student.date_graduated.strftime("%B %-d, %Y").upper()} PER REFERENDUM NO. {transcript.ref_no}, S. {transcript.series}
                          OF THE BOARD OF REGENTS, BICOL UNIVERSITY, LEGAZPI CITY.'''
        ws[f'A{i}'].alignment = opx.styles.Alignment(wrap_text=True, horizontal='center', vertical='center')

    if page_count > 1:
        for x in range(1, page_count):
            ws[f'I{x * 66 + 9}'] = f'{student.last_name}, {student.first_name} {student.middle_name}'.upper()
            ws[f'I{x * 66 + 10}'] = f'continuation of page {x}'

    indorser_name = attestor.name
    if attestor.title: indorser_name += f', {attestor.title}'

    reviewer_name = reviewer.name
    if reviewer.title: reviewer_name += f', {reviewer.title}'

    registrar_name = registrar.name
    if registrar.title: registrar_name += f', {registrar.title}'

    for x in range(0, page_count):
        ws[f'A{x * 66 + 58}'] = indorser_name
        ws[f'A{x * 66 + 59}'] = attestor.position
        ws[f'F{x * 66 + 61}'] = reviewer_name
        ws[f'F{x * 66 + 62}'] = reviewer.position
        ws[f'J{x * 66 + 62}'] = registrar_name
        ws[f'J{x * 66 + 63}'] = registrar.position
        ws[f'M{x * 66 + 65}'] = f'Revision: {transcript.revision}'

    ws[f'B{page_count * 66 + 121}'] = transcript.remarks

    file = f'applications/ors/documents/tor/{file_name}'
    wb.save(filename=file)

    from gluon.contenttype import contenttype
    response.headers['Content-Type'] = contenttype('xlsx')
    response.headers['Content-disposition'] = f'attachment; filename={file_name}'
    response.stream(file)

def tor():
    # controller for TOR view

    student_id = request.args(0)
    student = db(db.student.student_id == student_id).select().first()
    college = db(db.college.id == student.college_id).select().first()
    program = db(db.program.id == student.program_id).select().first()
    specialization = db(db.specialization.id == student.specialization_id).select().first() if student.specialization_id else None
    grades = db(db.grade.student_id == student.id).select(join=db.course.on(db.grade.course_id == db.course.id),
                                                          orderby=db.grade.term_year | db.grade.term_sem)

    terms = set()       # set of terms; terms are of type tuple: (term_sem, term_year); to be used as keys for term_data dictionary

    # store to terms set all terms attended by student
    for grade in grades:
        terms.add((grade.grade.term_sem, grade.grade.term_year))

    # sort the terms by school year
    terms = sorted(terms, key=lambda tup: tup[0])
    terms = sorted(terms, key=lambda tup: tup[1])

    term_data = {}      # dictionary of grades data per term: {term: {"grades": [], "units": int, "equivalent": float}}
    summary = {"units": 0, "equivalent": 0}     # summary data

    # iterate over the terms
    for term in terms:
        # initialize the contents of term_data
        term_data[term] = {"grades": [], "units": 0, "equivalent": 0}

        # check for each grade and store to term_data[term] if matches the term
        for grade in grades:
            if grade.grade.term_sem == term[0] and grade.grade.term_year == term[1]:
                term_data[term]["grades"].append(grade)
                term_data[term]["units"] += grade.course.units
                term_data[term]["equivalent"] += grade.grade.equivalent

        term_data[term]["equivalent"] = round(
            term_data[term]["equivalent"], 1)     # round term equivalent
        # add term units to summary
        summary["units"] += term_data[term]["units"]
        # add term equivalent to summary
        summary["equivalent"] += term_data[term]["equivalent"]

    summary["equivalent"] = round(summary["equivalent"], 1)
    summary["GWA"] = round(summary["equivalent"] / summary["units"], 4)     # calculate for GWA

    transcript = db(db.transcript.student_id == student.id).select().first()
    committee = db(db.committee.transcript_id == transcript.id).select(join=db.faculty.on(db.committee.faculty_id == db.faculty.id))

    committee_list = []

    # sort the committee members according to position hierarchy
    for member in committee:
        if member.committee.position.lower() == "chairperson":
            chair = member
        elif member.committee.position.lower() == "vice-chairperson":
            vice_chair = member
        else:
            committee_list.append(member)
    try:
        # append the vice-chair and chair to the committee_list respectively
        committee_list.extend([vice_chair, chair])
    except:
        pass

    attestor = db(db.faculty.id == transcript.attestor_id).select().first()
    reviewer = db(db.faculty.id == transcript.reviewer_id).select().first()
    registrar = db(db.staff.id == transcript.registrar_id).select().first()

    return locals()

def rle_record():
    # This function takes the student id as argument and returns the view for the student's RLE record

    # get the student id
    student_id = request.args(0)
    if not student_id:
        raise HTTP(400, "Bad request")

    # get data of student from database
    student = db(db.student.student_id == student_id).select().first()
    rle_courses = db().select(db.rle_course.ALL)
    attended_cr = db(db.attended_community_resource.student_id == student.id).select()
    rle_record = db(db.rle_record.student_id == student.id).select().first()

    # list the terms (year levels)
    terms = {1, 2, 3, 4}

    # initialize an empty dictionary for the list of courses for each term
    term_courses = {}

    # initialize the total values to zero
    total = {"lecture_total": 0, "lecture_units": 0, "rle_hours": 0, "rle_units": 0}

    # iterate over each rle course and store it to respective term and calculate for the total values
    for term in terms:
        term_courses[term] = []
        for course in rle_courses:
            if course.year_level == term:
                term_courses[term].append(course)
                total["lecture_total"] += course.lecture_total
                total["lecture_units"] += course.lecture_units
                total["rle_hours"] += course.rle_hours
                total["rle_units"] += course.rle_units

    # create a dictionary for the attended community resources of the student using the category as key
    # key 1: Barangay/Municipalities
    # key 2: Hospital/Clinics
    attended_cr_dict = {1: [], 2: []}

    # loop over the community resources and store to attended_cr_dict according to category
    for cr in attended_cr:
        if cr.community_resource_id.type == 1:
            attended_cr_dict[1].append(cr)
        elif cr.community_resource_id.type == 2:
            attended_cr_dict[2].append(cr)

    if not request.args(1):
        return locals()
    elif request.args(1) == "download":
        # script for generating xlsx file of TOR
        import openpyxl as opx

        # store the file name
        file_name = f'{student.last_name}-{student.first_name}-{student.middle_name}-RLE_Record.xlsx'.replace(" ", "_")

        # load template file
        wb = opx.load_workbook(filename='applications/ors/static/templates/RLE-Record-Template.xlsx')

        # get current worksheet
        ws = wb.active

        # write student data to file
        ws['B8'] = f'{student.last_name}, {student.first_name} {student.middle_name}'.upper()
        ws['I8'] = student.class_year

        # initialize row counter
        row_counter = 12

        # write courses to file
        for term in terms:
            ws[f'A{row_counter}'] = f'Ladder {term}'
            for course in term_courses[term]:
                ws[f'B{row_counter}'] = course.code_subject
                ws[f'C{row_counter}'] = course.code_digit
                ws[f'D{row_counter}'] = "-"
                ws.merge_cells(f'E{row_counter}:G{row_counter}')
                ws[f'E{row_counter}'] = course.title
                ws[f'H{row_counter}'] = course.lecture_total
                ws[f'I{row_counter}'] = course.lecture_units
                ws[f'J{row_counter}'] = course.rle_hours
                ws[f'K{row_counter}'] = course.rle_units

                # format cells font and alignment
                for row in ws.iter_cols(min_row=row_counter, min_col=1, max_row=row_counter, max_col=11):
                    for cell in row:
                        cell.font = opx.styles.Font(name="Times New Roman", size=10)

                for row in ws.iter_cols(min_row=row_counter, min_col=8, max_row=row_counter, max_col=11):
                    for cell in row:
                        cell.alignment = opx.styles.Alignment(horizontal="center")

                row_counter += 1

            row_counter += 1

        # write the total values to file
        ws[f'G{row_counter}'] = "TOTAL"
        ws[f'G{row_counter}'].alignment = opx.styles.Alignment(horizontal="right")
        ws[f'G{row_counter}'].font = opx.styles.Font(name="Times New Roman", size=10, bold=True)

        ws[f'H{row_counter}'] = total["lecture_total"]
        ws[f'I{row_counter}'] = total["lecture_units"]
        ws[f'J{row_counter}'] = f'{total["rle_hours"]:,}'
        ws[f'K{row_counter}'] = total["rle_units"]

        # format cells font, alignment, and border
        for row in ws.iter_cols(min_row=row_counter, min_col=8, max_row=row_counter, max_col=11):
            for cell in row:
                cell.font = opx.styles.Font(name="Times New Roman", size=10, bold=True)
                cell.alignment = opx.styles.Alignment(horizontal="center")
                cell.border = opx.styles.Border(bottom=opx.styles.Side(border_style="double"))

        # write attended community resources to file
        row_counter += 2
        ws.merge_cells(f'A{row_counter}:D{row_counter}')
        ws[f'A{row_counter}'] = "COMMUNITY RESOURCES"
        ws[f'A{row_counter}'].font = opx.styles.Font(name="Times New Roman", bold=True)

        row_counter += 1
        # loop for every category in attended_cr_dict
        for category in range(1, 3):
            # check if there are community resources for the current category
            if attended_cr_dict[category]:
                # write the category to file
                ws.merge_cells(f'B{row_counter}:G{row_counter}')
                if category == 1:
                    ws[f'B{row_counter}'] = "1. BARANGAY/MUNICIPALITIES"
                elif category == 2 and attended_cr_dict[1]:
                    ws[f'B{row_counter}'] = "2. HOSPITALS/CLINICS"
                elif category == 2 and not attended_cr_dict[1]:
                    ws[f'B{row_counter}'] = "1. HOSPITALS/CLINICS"
                # format the cell font
                ws[f'B{row_counter}'].font = opx.styles.Font(name="Times New Roman", size=10, bold=True)

                # write to file the second table header depending on the category
                ws.merge_cells(f'H{row_counter}:I{row_counter}')
                if category == 1:
                    ws[f'H{row_counter}'] = "No. of Families"
                elif category == 2 and attended_cr_dict[1]:
                    ws[f'H{row_counter}'] = "Daily Average Patient"
                # format the cell font and alignment
                ws[f'H{row_counter}'].font = opx.styles.Font(name="Times New Roman", size=10, bold=True)
                ws[f'H{row_counter}'].alignment = opx.styles.Alignment(horizontal="center")

                # write the third table header to file
                ws.merge_cells(f'J{row_counter}:K{row_counter}')
                ws[f'J{row_counter}'] = "Year Level"
                # format the cell font and alignment
                ws[f'J{row_counter}'].font = opx.styles.Font(name="Times New Roman", size=10, bold=True)
                ws[f'J{row_counter}'].alignment = opx.styles.Alignment(horizontal="center")

                row_counter += 1
                # loop over the community resources stored in the list for the current category
                for cr in attended_cr_dict[category]:
                    # write to file the name of the community resource and format cell alignment
                    ws.merge_cells(f'C{row_counter}:G{row_counter}')
                    ws[f'C{row_counter}'] = cr.community_resource_id.name
                    ws[f'C{row_counter}'].alignment = opx.styles.Alignment(wrap_text=True)
                    if len(cr.community_resource_id.name) > 57:
                        ws.row_dimensions[row_counter].height = 28

                    # write to file the community resource's total famiies or daily average patient depending on the category
                    ws.merge_cells(f'H{row_counter}:I{row_counter}')
                    ws[f'H{row_counter}'].alignment = opx.styles.Alignment(horizontal="center")
                    if category == 1:
                        ws[f'H{row_counter}'] = f'{cr.community_resource_id.families_total:,}'
                    elif category == 2:
                        ws[f'H{row_counter}'] = f'{cr.community_resource_id.patients_daily_ave:,}'

                    # write to file the year level when the student attended the community resource and format cell alignment
                    ws.merge_cells(f'J{row_counter}:K{row_counter}')
                    ws[f'J{row_counter}'] = cr.year_level
                    ws[f'J{row_counter}'].alignment = opx.styles.Alignment(horizontal="center")

                    # format cells font
                    for row in ws.iter_cols(min_row=row_counter, min_col=3, max_row=row_counter, max_col=11):
                        for cell in row:
                            cell.font = opx.styles.Font(name="Times New Roman", size=10)

                    row_counter += 1

        # write to file the date completed field
        row_counter += 1
        ws.merge_cells(f'A{row_counter}:B{row_counter}')
        ws[f'A{row_counter}'] = "Date Completed: "
        ws[f'A{row_counter}'].font = opx.styles.Font(name="Times New Roman", bold=True)
        ws[f'C{row_counter}'] = student.date_graduated.strftime("%B %-d, %Y")
        ws[f'C{row_counter}'].font = opx.styles.Font(name="Times New Roman")

        # write to file the name of the dean of the college and format cells font and alignment
        row_counter += 3
        ws.merge_cells(f'G{row_counter}:K{row_counter}')
        ws[f'G{row_counter}'] = f'{student.college_id.dean.name.upper()}{f", {student.college_id.dean.title}" if student.college_id.dean.title else ""}'
        ws[f'G{row_counter}'].font = opx.styles.Font(name="Times New Roman", bold=True)
        ws[f'G{row_counter}'].alignment = opx.styles.Alignment(horizontal="center")
        row_counter+=1
        ws.merge_cells(f'G{row_counter}:K{row_counter}')
        ws[f'G{row_counter}'] = "Dean"
        ws[f'G{row_counter}'].font = opx.styles.Font(name="Times New Roman")
        ws[f'G{row_counter}'].alignment = opx.styles.Alignment(horizontal="center")


        # write to file the name of the refistrar of the college and format cells font and alignment
        row_counter += 2
        ws[f'B{row_counter}'] = "Noted:"
        ws[f'B{row_counter}'].font = opx.styles.Font(name="Times New Roman")
        row_counter += 2
        ws.merge_cells(f'B{row_counter}:F{row_counter}')
        ws[f'B{row_counter}'] = f'{rle_record.registrar_id.name.upper()}{f", {rle_record.registrar_id.title}" if rle_record.registrar_id.title else ""}'
        ws[f'B{row_counter}'].font = opx.styles.Font(name="Times New Roman", bold=True)
        ws[f'B{row_counter}'].alignment = opx.styles.Alignment(horizontal="center")
        row_counter += 1
        ws.merge_cells(f'B{row_counter}:F{row_counter}')
        ws[f'B{row_counter}'] = f'{rle_record.registrar_id.position}'
        ws[f'B{row_counter}'].font = opx.styles.Font(name="Times New Roman")
        ws[f'B{row_counter}'].alignment = opx.styles.Alignment(horizontal="center")

        # write to file the revision number of the certificate
        ws['J58'] = f'Revision: {rle_record.revision}'

        # save the xlsx file
        file = f'applications/ors/documents/rle-record/{file_name}'
        wb.save(filename=file)

        # return the generated file as a response
        from gluon.contenttype import contenttype
        response.headers['Content-Type'] = contenttype('xlsx')
        response.headers['Content-disposition'] = f'attachment; filename={file_name}'
        response.stream(file)

def enrollment_certificate():
    student_id = request.args(0)
    if not student_id:
        raise HTTP(400, "Bad request")
        return

    #  get data of student from database
    student = db(db.student.student_id == student_id).select().first()
    enrollment_certificate = db(db.enrollment_certificate.id == student.id).select().first()

    registrar = db(db.staff.id).select().first()
    revision = db(db.enrollment_certificate.revision).select().first()

    honorific = {"Male": "Mr.", "Female": "Ms."}
    suffix = {1: "st", 2: "nd", 3: "rd", 4: "th", 5: "th", 6: "th"}
    day_issued = make_ordinal(enrollment_certificate.date_issued.strftime("%d"))

    if not request.args(1):
        return locals()

    elif request.args(1) == "download":
        # GENERATING THE XLSX FILE

        from openpyxl import load_workbook
        from openpyxl.styles import Font
        wb = load_workbook(filename='applications/ors/static/templates/Enrollment-Certificate-Template.xlsx')
        file_name = f'{student.last_name}-{student.first_name}-{student.middle_name}-Enrollment_Certificate.xlsx'.replace(" ", "_")

        sh1 = wb['Sheet1']
        b = Font(bold=True)
        # fsize = Font(size=15)
        sh1.cell(row=4, column=4, value='Republic of the Philippines')
        sh1.cell(row=5, column=4, value='Bicol University')
        sh1.cell(row=6, column=4, value=student.college_id.name)
        sh1.cell(row=7, column=4, value=student.college_id.address)
        sh1.cell(row=7, column=4, value='Tel. No. 'f'{student.college_id.contact_number}')
        sh1.cell(row=6, column=4).font = b

        sh1.cell(row=20, column=2, value='\tThis is to certify that ' + honorific[student.gender]+ ' ' + student.first_name+ ', '+ student.middle_name + ' '+ student.last_name+ ' a bona fied '+ student.year_level+ ' '+student.program_id.name+ ' major in ' + student.specialization_id.name +
        ' student of this college and is officially enrolled this' + ' ' + enrollment_certificate.term_sem  + ' Semester School Year ' +enrollment_certificate.term_year )
        sh1.cell(row=20, column=2).font = b
        sh1.cell(row=25, column=2, value='\tIssued this ' + enrollment_certificate.date_issued.strftime("%dsuffix day of %B, %Y").replace('suffix', str(suffix_generator(enrollment_certificate.date_issued.day)) ) + ' for reference purposes.' )

        sh1.cell(row=29, column=6, value=f'{registrar.name}'.upper())
        sh1.cell(row=29, column=6).font = b
        sh1.cell(row=30, column=6, value=f'{registrar.position}')

        sh1.cell(row=39, column=8, value='Revision:'f'{enrollment_certificate.revision}')

        file = f'applications/ors/documents/enrollment-certificate/{file_name}'
        wb.save(filename=file)

        # return the generated file as a response
        from gluon.contenttype import contenttype
        response.headers['Content-Type'] = contenttype('xlsx')
        response.headers['Content-disposition'] = f'attachment; filename={file_name}'
        response.stream(file)


def good_moral_certificate():
    student_id = request.args(0)
    if not student_id:
        raise HTTP(400, "Bad request")

    student = db(db.student.student_id == student_id).select().first()
    good_moral_certificate = db(db.good_moral_certificate.student_id == student.id).select().first()

    honorific = {"Male": "Mr.", "Female": "Ms."}
    suffix = {1: "st", 2: "nd", 3: "rd", 4: "th", 5: "th", 6: "th"}
    pronoun_sub = {"Male": "he", "Female": "she"}
    pronoun_pos = {"Male": "his", "Female": "her"}
    day_issued = make_ordinal(good_moral_certificate.date_issued.strftime("%d"))

    return locals()

def grades_certificate():
    # get the student id
    student_id = request.args(0)
    if not student_id:
        raise HTTP(400, "Bad request")

    student = db(db.student.student_id==student_id).select().first()
    college = db(db.college.id==student.college_id).select().first()
    program = db(db.program.id==student.program_id).select().first()
    grades = db(db.grade.student_id==student.id).select(
        join=db.course.on(db.grade.course_id==db.course.id),
        orderby=db.grade.term_year|db.grade.term_sem
    )
    college_registrar = db(student.id==db.grades_certificate.student_id).select().first()
    registrar = college_registrar.registrar

    terms = set()       # set of terms; terms are of type tuple: (term_sem, term_year); to be used as keys for term_data dictionary

    # store to terms set all terms attended by student
    for grade in grades:
        terms.add((grade.grade.term_sem, grade.grade.term_year))

    # sort the terms by school year
    terms = sorted(terms, key=lambda tup: tup[0])
    terms = sorted(terms, key=lambda tup: tup[1])

    term_data = {}      # dictionary of grades data per term: {term: {"grades": [], "units": int, "equivalent": float}}
    summary = {"units": 0, "equivalent": 0}     # summary data

    # iterate over the terms
    for term in terms:
        term_data[term] = {"grades": [], "units": 0, "equivalent": 0}       # initialize the contents of term_data

        # check for each grade and store to term_data[term] if matches the term
        for grade in grades:
            if grade.grade.term_sem == term[0] and grade.grade.term_year == term[1]:
                term_data[term]["grades"].append(grade)
                term_data[term]["units"] += grade.course.units


        summary["units"] += term_data[term]["units"]

    if not request.args(1):
        return locals()

    elif request.args(1) == "download":
        # GENERATING THE XLSX FILE

        from openpyxl import load_workbook
        wb = load_workbook(filename='applications/ors/static/templates/COG-Template.xlsx')
        file_name = f'{student.last_name}-{student.first_name}-{student.middle_name}-COG.xlsx'.replace(" ", "_")

        sh1 = wb['Sheet1']
        sh1.cell(row=6,column=4, value=f'{college.name}'.upper())
        sh1.cell(row=7,column=4, value=college.address)
        sh1.cell(row=8,column=4, value=college.contact_number)

        sh1.cell(row=20, column=2, value='\tOn the basis of records on file in this office, I hereby certify that '+ student.first_name+ ', '+ student.middle_name + ' '+ student.last_name+ ', a '+ student.year_level+ ' '+program.name+' student of the college, took the following subjects with the corresponding grades and units, to wit:')
        i = 25

        for term in terms:

            j = 2
            grades = term_data[term]["grades"]

            sh1.cell(row=i,column=j, value=f'{term[0]}')
            sh1.cell(row=i+1,column=j, value=f'{term[1]}')
            j = j+1

            for grade in grades:

                sh1.cell(row =i, column=j, value=f'{grade.course.code}')
                j = j+1

                sh1.cell(row =i, column=j, value=f'{grade.course.title}')
                j = j+1

                if grade.grade.final_grade != 4.0:
                    sh1.cell(row =i, column=j, value=f'{grade.grade.final_grade}')
                else:
                    sh1.cell(row =i, column=j, value='INC')
                j = j+1

                if grade.grade.final_grade == 4.0:
                    sh1.cell(row =i, column=j, value=f'{grade.grade.removal_rating}')
                j = j+1

                sh1.cell(row =i, column=j, value=f'{grade.course.units}')

                j = 3
                i = i+1


        sh1.cell(row=41,column=2, value='\tIssued this ' +college_registrar.date_issued.strftime("%dsuffix day of %B, %Y").replace('suffix', str(suffix_generator(college_registrar.date_issued.day)) ) + ' for reference purposes.' )

        sh1.cell(row =43, column=6, value=f'{registrar.name}'.upper())

        file = f'applications/ors/documents/cog/{file_name}'
        wb.save(filename=file)

        # return the generated file as a response
        from gluon.contenttype import contenttype
        response.headers['Content-Type'] = contenttype('xlsx')
        response.headers['Content-disposition'] = f'attachment; filename={file_name}'
        response.stream(file)

def suffix_generator(d):
    return 'th' if 11<=d<=13 else {1:'st',2:'nd',3:'rd'}.get(d%10, 'th')
