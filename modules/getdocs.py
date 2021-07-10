import openpyxl as opx
import pydal
from os import path

def load_workbook(wb_path):
    if path.exists(wb_path):
        return opx.load_workbook(wb_path)
    return opx.Workbook()

def generate_tor_file(student_id):
    student_id = request.args(0)
    student = db(db.student.student_id == student_id).select().first()
    college = db(db.college.id == student.college_id).select().first()
    program = db(db.program.id == student.program_id).select().first()

    if student.specialization_id:
        specialization = db(db.specialization.id == student.specialization_id).select().first()
    else:
        specialization = None
    
    grades = db(db.grade.student_id == student.id).select(join=db.course.on(db.grade.course_id == db.course.id),
                                                          orderby=db.grade.term_year | db.grade.term_sem)

    terms = set()       # set of terms; terms are of type tuple: (term_sem, term_year); to be used as keys for term_data dictionary

    # store to terms set all terms attended by student
    for grade in grades:
        terms.add((grade.grade.term_sem, grade.grade.term_year))

    # sort the terms by school year
    terms = sorted(terms, key=lambda tup: tup[0])
    terms = sorted(terms, key=lambda tup: tup[1])

    term_data = {}      # dictionary of grades data per term: {term: {"grades": [], "units": int, "equivalent": float}}
    summary = {"units": 0, "equivalent": 0}     # summary data

    # iterate over the terms
    for term in terms:
        # initialize the contents of term_data
        term_data[term] = {"grades": [], "units": 0, "equivalent": 0}

        # check for each grade and store to term_data[term] if matches the term
        for grade in grades:
            if grade.grade.term_sem == term[0] and grade.grade.term_year == term[1]:
                term_data[term]["grades"].append(grade)
                term_data[term]["units"] += grade.course.units
                term_data[term]["equivalent"] += grade.grade.equivalent

        term_data[term]["equivalent"] = round(
            term_data[term]["equivalent"], 1)     # round term equivalent
        # add term units to summary
        summary["units"] += term_data[term]["units"]
        # add term equivalent to summary
        summary["equivalent"] += term_data[term]["equivalent"]

    summary["equivalent"] = round(summary["equivalent"], 1)
    summary["GWA"] = round(summary["equivalent"] /
                           summary["units"], 4)     # calculate for GWA

    transcript = db(db.transcript.student_id == student.id).select().first()
    committee = db(db.committee.transcript_id == transcript.id).select(
        join=db.faculty.on(db.committee.faculty_id == db.faculty.id))

    committee_list = []

    # sort the committee members according to position hierarchy
    for member in committee:
        if member.committee.position.lower() == "chairperson":
            chair = member
        elif member.committee.position.lower() == "vice-chairperson":
            vice_chair = member
        else:
            committee_list.append(member)
    try:
        # append the vice-chair and chair to the committee_list respectively
        committee_list.extend([vice_chair, chair])
    except:
        pass

    attestor = db(db.faculty.id == transcript.attestor_id).select().first()
    reviewer = db(db.faculty.id == transcript.reviewer_id).select().first()
    registrar = db(db.staff.id == transcript.registrar_id).select().first()


    file_path = f'{student.last_name}-{student.first_name}-{student.middle_name}-TOR.xlsx'.replace(" ", "_")